#!/usr/bin/env python3
"""
KTTA 標準 Excel 組合せ表パーサー
====================================
釧路卓球協会で実際に使われている形式の Excel を読み取る専用パーサー。

形式の特徴:
  ・1シートに ○種目名 でセクション分割
  ・LEFT 半分: A列=position, B列=選手名 (B29:B30 merged), C/D列=所属
  ・MIDDLE: S列=選手名 (S29:T30 merged), U列=所属  ← 右半分ブラケット
  ・RIGHT 列リスト: V列=position番号 (V29:V30 merged), X列=所属, Y列=選手名
  ・団体戦は別レイアウト (チーム名+メンバー)

実行:
    python3 parse_ktta_bracket.py FILE.xlsx [--format singles|doubles|team]
                                            [--event "種目名"]
                                            [--sheet 名前]
                                            [--all-sheets]

出力 JSON 形式 (db.importBracket と互換):
    {
      "format": "tabletennis-seed-list-v1",
      "event": "一般男子シングルス",
      "regenerate": true,
      "players": [
        { "name": "木村 翔", "team": "AMATAKU", "seed": 1 },
        ...
      ]
    }

または bracket-tree 形式:
    {
      "format": "tabletennis-tournament-v1",
      "brackets": [
        { "event": "...", "players": [...] }
      ]
    }
"""
from __future__ import annotations
import openpyxl
from openpyxl.utils import get_column_letter
import json
import sys
import re
import argparse
from collections import defaultdict
from pathlib import Path


# 種目ヘッダー検出パターン
SECTION_HEADER_RE = re.compile(r'^\s*[○◯◎●]\s*(.+?)\s*$')

# 種目名カテゴリ
TEAM_KEYWORDS = ("団体", "チーム")
DOUBLES_KEYWORDS = ("ダブルス", "ミックス", "混合", "ペア")
# それ以外は singles


def detect_format(event_name: str, hint: str | None) -> str:
    """種目名と --format ヒントから format を確定"""
    if hint in ("singles", "doubles", "team"):
        return hint
    if event_name:
        for kw in TEAM_KEYWORDS:
            if kw in event_name:
                return "team"
        for kw in DOUBLES_KEYWORDS:
            if kw in event_name:
                return "doubles"
    return "singles"


def strip_parens(s: str) -> str:
    """『(AMATAKU)』→『AMATAKU』 / 『（AMATAKU）』→『AMATAKU』"""
    if s is None:
        return ""
    s = str(s).strip()
    m = re.match(r'^[(（]\s*(.*?)\s*[)）]$', s)
    return m.group(1) if m else s


def normalize_name(s: str) -> str:
    if not s:
        return ""
    s = str(s).strip()
    # 全角空白→半角・連続→単一
    s = s.replace("　", " ")
    s = re.sub(r'\s+', " ", s).strip()
    # 末尾の「君」「さん」「ちゃん」「選手」を除去
    s = re.sub(r'(君|くん|さん|ちゃん|選手|様)$', "", s).strip()
    return s


def is_label_like(s: str) -> bool:
    """『氏名』『所属』など、データではないラベルか"""
    if not s:
        return True
    s = str(s).strip()
    LABELS = {"氏名", "所属", "選手名", "団体名", "チーム名", "メンバー",
              "代表者", "選手", "ペア", "ダブルス", "シングルス",
              "決勝", "準決勝", "準々決勝", "ベスト16", "ベスト32",
              "BYE", "bye", "不戦勝", "棄権"}
    if s in LABELS:
        return True
    # ※/★/●で始まる注意書き
    if s.startswith(("※", "★", "●", "◯", "○", "・", "■", "□")):
        return True
    return False


def find_sections(ws):
    """シート内の ○種目 ヘッダー位置を全て見つけて返す。
       戻り値: [(start_row, event_name), ...]"""
    sections = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(15, ws.max_column + 1)):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            m = SECTION_HEADER_RE.match(str(v))
            if m:
                sections.append((r, m.group(1).strip()))
                break
    # セクション末尾を補完 (次のセクション開始 -1 までを範囲とする)
    result = []
    for i, (sr, name) in enumerate(sections):
        end = sections[i + 1][0] - 1 if i + 1 < len(sections) else ws.max_row
        result.append({"start": sr, "end": end, "event": name})
    return result


def get_merged_value(ws, row, col):
    """セル値を取得。merged の場合は左上の値を返す"""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    # merged check
    for mr in ws.merged_cells.ranges:
        if (mr.min_row <= row <= mr.max_row and
                mr.min_col <= col <= mr.max_col):
            return ws.cell(row=mr.min_row, column=mr.min_col).value
    return None


def is_top_of_merge(ws, row, col):
    """このセルが merged の最上行か (or merged でないか)"""
    for mr in ws.merged_cells.ranges:
        if (mr.min_row <= row <= mr.max_row and
                mr.min_col <= col <= mr.max_col):
            return row == mr.min_row
    return True


# ─────────────────────────────────────────────────────
# シングルス: 左右両側のシード一覧を読み取る
# ─────────────────────────────────────────────────────
def parse_singles_section(ws, section, verbose=False):
    """LEFT (cols A=position, B=name, C=team) と
       RIGHT (cols V=position, X=team, Y=name) から選手を抽出"""
    players = []
    sr, er = section["start"], section["end"]
    event_name = section["event"]

    # ─── LEFT 側: A列の数値を見て name+team を取る ───
    for r in range(sr + 1, er + 1):
        a_val = ws.cell(row=r, column=1).value  # A
        if a_val is None or not isinstance(a_val, (int, float)):
            continue
        # B列が merged の上端でなければスキップ (重複防止)
        if not is_top_of_merge(ws, r, 2):
            continue
        b_val = get_merged_value(ws, r, 2)  # B = name
        c_val = get_merged_value(ws, r, 3)  # C = team (in parens)
        if c_val is None:
            c_val = get_merged_value(ws, r, 4)  # D fallback
        name = normalize_name(b_val)
        team = strip_parens(c_val) if c_val else ""
        if name and not is_label_like(name):
            players.append({
                "name": name,
                "team": team,
                "seed": int(a_val),
                "_side": "L",
            })

    # ─── MIDDLE 側 (中央=右半分ブラケットを表示): S=name, U=team ───
    # V列の数値がついた行で読む。複数行 merged。
    for r in range(sr + 1, er + 1):
        v_val = ws.cell(row=r, column=22).value  # V
        if v_val is None or not isinstance(v_val, (int, float)):
            continue
        if not is_top_of_merge(ws, r, 19):  # S
            continue
        s_val = get_merged_value(ws, r, 19)  # S = name
        u_val = get_merged_value(ws, r, 21)  # U = team (in parens)
        name = normalize_name(s_val)
        team = strip_parens(u_val) if u_val else ""
        if name and not is_label_like(name):
            # 既に LEFT 側にいないかチェック (重複防止)
            existing = next((p for p in players if p["name"] == name and p["team"] == team), None)
            if not existing:
                players.append({
                    "name": name,
                    "team": team,
                    "seed": int(v_val),
                    "_side": "R",
                })

    # ─── RIGHT エントリーリスト (V列=position): X=team, Y=name ───
    # これは LEFT/MIDDLE で取れなかった選手を補う。または完全な seed list。
    for r in range(sr + 1, er + 1):
        v_val = ws.cell(row=r, column=22).value
        y_val = ws.cell(row=r, column=25).value
        x_val = ws.cell(row=r, column=24).value
        name = normalize_name(y_val)
        team = strip_parens(x_val) if x_val else ""
        if not name or is_label_like(name):
            continue
        # 既存と重複してたら追加しない
        existing = next((p for p in players if p["name"] == name and p["team"] == team), None)
        if existing:
            continue
        # V列の数値が無くても、Y/X 列に名前があれば登録 (seed 不明=0)
        seed = int(v_val) if isinstance(v_val, (int, float)) else 0
        players.append({
            "name": name,
            "team": team,
            "seed": seed,
            "_side": "R" if seed >= 12 else "L",
        })

    # _side フィールドを削除
    for p in players:
        p.pop("_side", None)
    # seed 順
    players.sort(key=lambda p: (p["seed"] or 9999, p["name"]))
    return players


# ─────────────────────────────────────────────────────
# ダブルス: 1ペア = 2名がペア化される
# ─────────────────────────────────────────────────────
def parse_doubles_section(ws, section, verbose=False):
    """シングルスと同様だが、各位置に 2名のペアが入る。
       行2行で1ペアを表現する場合 (B行=選手1, C行=選手2) や
       同行 列を分けて表現する場合がある。
       簡易実装: name1 / name2 と検出された場合はペアとして扱う"""
    players = []
    sr, er = section["start"], section["end"]
    event_name = section["event"]

    # LEFT
    for r in range(sr + 1, er + 1, 1):
        a_val = ws.cell(row=r, column=1).value
        if a_val is None or not isinstance(a_val, (int, float)):
            continue
        # ペアの場合、B列とC列に2名 (異なる行 or 列)
        b1 = get_merged_value(ws, r, 2) or ""
        b2 = get_merged_value(ws, r + 1, 2) or ""
        c_val = get_merged_value(ws, r, 3) or get_merged_value(ws, r, 4) or ""
        n1 = normalize_name(b1)
        n2 = normalize_name(b2)
        team = strip_parens(c_val) if c_val else ""
        pair_name = (n1 + "/" + n2) if (n1 and n2 and n1 != n2) else (n1 or n2)
        if pair_name and not is_label_like(pair_name):
            players.append({
                "name": pair_name,
                "name1": n1,
                "name2": n2,
                "team": team,
                "is_doubles": True,
                "seed": int(a_val),
            })

    # RIGHT エントリーリスト
    last_y_pair = []
    for r in range(sr + 1, er + 1):
        v_val = ws.cell(row=r, column=22).value
        y_val = ws.cell(row=r, column=25).value
        x_val = ws.cell(row=r, column=24).value
        name = normalize_name(y_val)
        team = strip_parens(x_val) if x_val else ""
        if not name or is_label_like(name):
            continue
        # 2行で1ペアと仮定 (V列の merged で確認)
        if isinstance(v_val, (int, float)):
            # 新しいペアの開始
            last_y_pair = [{"name": name, "team": team}]
        else:
            # 既存ペアに2人目を追加
            if last_y_pair and len(last_y_pair) == 1:
                last_y_pair.append({"name": name, "team": team})
                # commit
                p1, p2 = last_y_pair
                pair_name = p1["name"] + "/" + p2["name"]
                # 重複チェック
                existing = next((p for p in players if p["name"] == pair_name), None)
                if not existing:
                    players.append({
                        "name": pair_name,
                        "name1": p1["name"],
                        "name2": p2["name"],
                        "team": p1["team"],  # 同じ所属を仮定
                        "is_doubles": True,
                        "seed": 0,
                    })
                last_y_pair = []

    players.sort(key=lambda p: (p["seed"] or 9999, p["name"]))
    return players


# ─────────────────────────────────────────────────────
# 団体戦: 1チーム = 4〜6名のメンバー
# ─────────────────────────────────────────────────────
def parse_team_section(ws, section, verbose=False):
    """A列の数値ごとに 1チーム。
       B列=チーム名 (merged 数行)、C/D列=メンバー名 (複数行に並ぶ)
       戻り値: チーム単位の entrants"""
    teams = []
    sr, er = section["start"], section["end"]

    # 各 A 列の数値で区切る
    team_rows = []  # [(start_row, end_row, team_pos), ...]
    last_start = None
    last_pos = None
    for r in range(sr + 1, er + 2):
        a_val = ws.cell(row=r, column=1).value if r <= ws.max_row else None
        if isinstance(a_val, (int, float)):
            if last_start is not None:
                team_rows.append((last_start, r - 1, last_pos))
            last_start = r
            last_pos = int(a_val)
    if last_start is not None:
        team_rows.append((last_start, er, last_pos))

    for sr_, er_, pos in team_rows:
        # B列 (merged) でチーム名
        team_name = normalize_name(get_merged_value(ws, sr_, 2))
        if not team_name or is_label_like(team_name):
            continue
        # メンバー: C, D 列を行範囲で読み取る
        members = []
        for r in range(sr_, er_ + 1):
            for col in (3, 4):  # C, D
                v = ws.cell(row=r, column=col).value
                nm = normalize_name(v)
                if nm and not is_label_like(nm):
                    if nm not in [m["name"] for m in members]:
                        members.append({"name": nm, "team": team_name})
        teams.append({
            "name": team_name,
            "team": team_name,  # 団体戦なので名前=チーム名
            "seed": pos,
            "is_team": True,
            "members": members,
        })

    # team を seed 順
    teams.sort(key=lambda t: (t["seed"] or 9999, t["name"]))
    # entrants 形式に変換: 1チーム = 1エントリー (メンバー list 付き)
    players = []
    for t in teams:
        players.append({
            "name": t["name"],
            "team": t["name"],
            "seed": t["seed"],
            "is_team": True,
            "members": [m["name"] for m in t["members"]],
        })
    return players


# ─────────────────────────────────────────────────────
# メイン解析
# ─────────────────────────────────────────────────────
def parse_sheet(ws, sheet_name, format_hint=None, event_hint=None, verbose=False):
    """シート1枚を解析。複数セクションあれば全て返す"""
    sections = find_sections(ws)
    if verbose:
        print(f"  Detected sections: {[s['event'] for s in sections]}", file=sys.stderr)

    # セクションが無ければ、シート全体を1セクションとして扱う
    if not sections:
        event_name = event_hint or sheet_name
        sections = [{"start": 0, "end": ws.max_row, "event": event_name}]

    brackets = []
    for sec in sections:
        ev_name = sec["event"]
        fmt = detect_format(ev_name, format_hint)
        if fmt == "team":
            players = parse_team_section(ws, sec, verbose)
        elif fmt == "doubles":
            players = parse_doubles_section(ws, sec, verbose)
        else:
            players = parse_singles_section(ws, sec, verbose)
        if not players:
            if verbose:
                print(f"  Section '{ev_name}' produced 0 players", file=sys.stderr)
            continue
        brackets.append({
            "format": "tabletennis-seed-list-v1",
            "event": ev_name,
            "type": fmt,
            "regenerate": True,
            "auto_link_to_players": False,
            "players": players,
        })
    return brackets


def parse_workbook(path, format_hint=None, event_hint=None,
                   sheet=None, all_sheets=False, verbose=False):
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet:
        sheets = [sheet] if sheet in wb.sheetnames else []
        if not sheets:
            return {"error": f"シート '{sheet}' が見つかりません",
                    "available_sheets": wb.sheetnames}
    elif all_sheets:
        sheets = wb.sheetnames
    else:
        sheets = [wb.sheetnames[0]]

    all_brackets = []
    for sn in sheets:
        ws = wb[sn]
        if verbose:
            print(f"=== Sheet: {sn} ===", file=sys.stderr)
        b = parse_sheet(ws, sn, format_hint, event_hint, verbose)
        all_brackets.extend(b)

    if not all_brackets:
        return {
            "error": "認識可能な種目セクションが見つかりませんでした",
            "hint": "シート内に『○種目名』のヘッダー行があることを確認してください",
            "available_sheets": wb.sheetnames,
        }

    # 単一セクションなら直接 seed-list 形式で返す (importBracket 互換)
    if len(all_brackets) == 1:
        return all_brackets[0]

    # 複数セクションなら tournament 形式
    return {
        "format": "tabletennis-tournament-v1",
        "tournament": {
            "name": Path(path).stem,
        },
        "brackets": all_brackets,
    }


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("file", help="Excel ファイル (.xlsx)")
    ap.add_argument("--format", choices=("singles", "doubles", "team", "auto"),
                    default="auto", help="種目形式 (default: 自動判定)")
    ap.add_argument("--event", help="種目名 ヒント (セクションヘッダーが無い場合)")
    ap.add_argument("--sheet", help="特定のシートのみ")
    ap.add_argument("--all-sheets", action="store_true", help="全シート")
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args()

    fh = None if args.format == "auto" else args.format
    try:
        result = parse_workbook(args.file, format_hint=fh,
                                event_hint=args.event,
                                sheet=args.sheet,
                                all_sheets=args.all_sheets,
                                verbose=args.verbose)
    except FileNotFoundError:
        print(json.dumps({"error": "ファイルが見つかりません"}), file=sys.stdout)
        sys.exit(1)
    except Exception as e:
        print(json.dumps({"error": f"解析エラー: {type(e).__name__}: {e}"}),
              file=sys.stdout)
        sys.exit(1)

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
