"""
grid.py — Excel ワークシートを「値 + 罫線」の正規化グリッドにする層。

openpyxl の生データには2つの厄介な点があり、ここで吸収する:

  1) 結合セル(merged cells): 値は左上セルにのみ入り、他は None。
     → value(r,c) は結合範囲を解決して左上の値を返す。
     → merge_origin(r,c) で「その結合ブロックの代表(左上)座標」を得る。

  2) 罫線の二重表現: セル(r,c)とその下(r+1,c)の境界線は、
     Excel上では「(r,c)の下罫線」または「(r+1,c)の上罫線」のどちらでも
     表現され得る(作成ツール依存)。同様に左右も (r,c)の右 / (r,c+1)の左。
     → edge_below(r,c) / edge_right(r,c) は両表現を OR して
        「2セル間に線があるか」を一意に判定する(罫線追跡の信頼性の要)。

座標は openpyxl に合わせ row/col とも 1 始まり。
"""

from __future__ import annotations
from dataclasses import dataclass
import openpyxl


# 「線あり」とみなす罫線スタイル。hair(極細)も実データで連結線に使われるため含める。
# none/None は線なし。
def _has_style(side) -> bool:
    return bool(side is not None and side.style)


@dataclass
class GridModel:
    sheet_name: str
    max_row: int
    max_col: int
    # 値: (row,col)->str（結合解決済み・前後空白除去・None は格納しない）
    _values: dict
    # 各セルの生の四辺罫線フラグ: (row,col)->(L,T,R,B) bool
    _borders: dict
    # 結合: (row,col)->(min_row,min_col,max_row,max_col)  範囲内の全セルが代表範囲を指す
    _merge_of: dict

    # ---- 値 ----
    def value(self, r: int, c: int):
        if r < 1 or c < 1 or r > self.max_row or c > self.max_col:
            return None
        rng = self._merge_of.get((r, c))
        if rng:
            return self._values.get((rng[0], rng[1]))
        return self._values.get((r, c))

    def raw_value(self, r: int, c: int):
        """結合解決せず、そのセル自身の値だけ返す(左上判定用)。"""
        return self._values.get((r, c))

    def merge_origin(self, r: int, c: int):
        rng = self._merge_of.get((r, c))
        return (rng[0], rng[1]) if rng else (r, c)

    def merge_range(self, r: int, c: int):
        """(min_row,min_col,max_row,max_col) or None"""
        return self._merge_of.get((r, c))

    def is_merge_origin(self, r: int, c: int) -> bool:
        rng = self._merge_of.get((r, c))
        return (rng is None) or (rng[0] == r and rng[1] == c)

    # ---- 生の四辺罫線 ----
    def _b(self, r: int, c: int):
        return self._borders.get((r, c), (False, False, False, False))

    def border_left(self, r, c):   return self._b(r, c)[0]
    def border_top(self, r, c):    return self._b(r, c)[1]
    def border_right(self, r, c):  return self._b(r, c)[2]
    def border_bottom(self, r, c): return self._b(r, c)[3]

    # ---- 正規化エッジ(2セル間に線があるか) ----
    def edge_below(self, r: int, c: int) -> bool:
        """(r,c) と (r+1,c) の間に水平線があるか。"""
        if self.border_bottom(r, c):
            return True
        if r + 1 <= self.max_row and self.border_top(r + 1, c):
            return True
        return False

    def edge_right(self, r: int, c: int) -> bool:
        """(r,c) と (r,c+1) の間に垂直線があるか。"""
        if self.border_right(r, c):
            return True
        if c + 1 <= self.max_col and self.border_left(r, c + 1):
            return True
        return False

    def edge_above(self, r: int, c: int) -> bool:
        return self.edge_below(r - 1, c) if r > 1 else self.border_top(r, c)

    def edge_left(self, r: int, c: int) -> bool:
        return self.edge_right(r, c - 1) if c > 1 else self.border_left(r, c)

    # ---- 走査補助 ----
    def iter_values(self):
        """(row,col,value) を行→列順で。結合は代表(左上)1回のみ。"""
        for (r, c), v in sorted(self._values.items()):
            yield r, c, v


def load_grid(path_or_ws, sheet_name: str | None = None) -> GridModel:
    """xlsx パス(+シート名) or 既に開いた worksheet から GridModel を構築。"""
    if isinstance(path_or_ws, str):
        wb = openpyxl.load_workbook(path_or_ws, data_only=True, read_only=False)
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    else:
        ws = path_or_ws

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    # 結合範囲マップ
    merge_of = {}
    for rng in ws.merged_cells.ranges:
        key = (rng.min_row, rng.min_col, rng.max_row, rng.max_col)
        for rr in range(rng.min_row, rng.max_row + 1):
            for cc in range(rng.min_col, rng.max_col + 1):
                merge_of[(rr, cc)] = key

    values = {}
    borders = {}
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            v = cell.value
            if v is not None:
                s = str(v).strip()
                if s != "":
                    values[(cell.row, cell.column)] = s
            b = cell.border
            L = _has_style(b.left); T = _has_style(b.top)
            R = _has_style(b.right); B = _has_style(b.bottom)
            if L or T or R or B:
                borders[(cell.row, cell.column)] = (L, T, R, B)

    return GridModel(
        sheet_name=getattr(ws, "title", sheet_name or ""),
        max_row=max_row, max_col=max_col,
        _values=values, _borders=borders, _merge_of=merge_of,
    )


def sheet_names(path: str) -> list:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()
