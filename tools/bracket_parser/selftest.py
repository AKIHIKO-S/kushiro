"""
selftest.py — 依存フリーの自己テスト(pytest不要)。

  python3 -m bracket_parser.selftest          # 合成フィクスチャ + 単体
  python3 -m bracket_parser.selftest --real    # 加えて実データ検証(あれば)

合成フィクスチャは openpyxl でその場生成する偽データ(PIIなし)。
実データ検証はローカルに VICTAS ファイルと正解JSONがある時のみ実行し、
無ければスキップ(公開リポジトリ/CIでも安全)。
"""

from __future__ import annotations
import sys
import os
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Border, Side

from . import tokens as T
from . import api, emit as M, topology as TP
from .grid import load_grid
from .entries import extract_leaves

_PASS = 0
_FAIL = 0


def check(cond, msg):
    global _PASS, _FAIL
    if cond:
        _PASS += 1
    else:
        _FAIL += 1
        print(f"  FAIL: {msg}")


def eq(a, b, msg):
    check(a == b, f"{msg}  (got {a!r}, want {b!r})")


# ──────────────────────────────────────────────────────────
# 1. tokens 単体
# ──────────────────────────────────────────────────────────
def test_tokens():
    print("[tokens]")
    eq(T.clean_team("(スマイルクラブ)"), "スマイルクラブ", "clean_team strips outer parens")
    eq(T.clean_team("ＡＳＴ根室"), "AST根室", "clean_team NFKC")
    eq(T.clean_team("A中(分校)"), "A中(分校)", "clean_team keeps inner parens")
    eq(T.strip_parens("山田 太郎 (A中)"), "山田 太郎", "strip_parens trailing note")
    eq(T.join_pair("山田 太郎", "鈴木 一郎"), "山田 太郎 / 鈴木 一郎", "join_pair")
    eq(T.join_pair("山田", "山田"), "山田", "join_pair dedup equal")
    eq(T.join_teams(["AMATAKU", "個人"]), "AMATAKU / 個人", "join_teams two")
    eq(T.join_teams(["スマイルクラブ"]), "スマイルクラブ", "join_teams one")
    eq(T.join_teams(["AMATAKU", "AMATAKU"]), "AMATAKU / AMATAKU", "join_teams keeps dup")
    eq(T.to_int("１２"), 12, "to_int full-width")
    eq(T.to_int("3."), 3, "to_int trailing dot")
    eq(T.to_int("abc"), None, "to_int non-number")
    check(T.is_bye("不戦"), "is_bye 不戦")
    check(T.is_bye(""), "is_bye empty")
    check(T.is_bye("―"), "is_bye dash")
    check(T.looks_like_name("田中 花子"), "looks_like_name ok")
    check(not T.looks_like_name("1回戦"), "looks_like_name rejects round label")
    check(not T.looks_like_name("123"), "looks_like_name rejects number")
    check(T.is_label_like("準決勝"), "is_label_like round word")
    check(T.is_label_like("2026.2.11"), "is_label_like date")


# ──────────────────────────────────────────────────────────
# 2. grid 合成: マージ解決 + 罫線OR
# ──────────────────────────────────────────────────────────
def test_grid():
    print("[grid]")
    wb = Workbook(); ws = wb.active
    ws["A1"] = "X"
    ws.merge_cells("A1:A2")              # 縦結合
    ws["C3"] = "v"
    thin = Side(style="thin")
    # B5 の下罫線(=下隣 B6 の上罫線として表現してもedge_belowはTrue)
    ws["B6"].border = Border(top=thin)
    # D5 の右罫線
    ws["D5"].border = Border(right=thin)
    # E5 の左罫線(=D5 の右としても edge_right True)
    ws["E5"].border = Border(left=thin)
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "g.xlsx"); wb.save(p)
        g = load_grid(p, ws.title)
        eq(g.value(2, 1), "X", "merge value resolves to top-left")
        eq(g.merge_origin(2, 1), (1, 1), "merge_origin")
        check(g.edge_below(5, 2), "edge_below via neighbor top border")
        check(g.edge_right(5, 4), "edge_right via own right border")
        check(g.edge_right(5, 4), "edge_right via neighbor left border (D/E)")


# ──────────────────────────────────────────────────────────
# 3. 合成ブラケット(偽データ): 各レイアウトの抽出
# ──────────────────────────────────────────────────────────
def _set(ws, r, c, v):
    ws.cell(row=r, column=c, value=v)


def _build_singles(ws):
    # 左半分: [num][name][team] 行ごと, 右半分: [team][name][num]
    # 左 1..6 @ rows 2,4,.. (col A=num,B=name,C=team)
    L = [("ア 太郎", "赤クラブ"), ("イ 次郎", "青中"), ("ウ 三郎", "緑高"),
         ("エ 四郎", "黄校"), ("オ 五郎", "紫団"), ("カ 六郎", "桃ク")]
    for i, (nm, tm) in enumerate(L):
        r = 2 + i * 2
        _set(ws, r, 1, i + 1)
        _set(ws, r, 2, nm)
        _set(ws, r, 3, f"({tm})")
    # 右 7..12 (col H=num, G=team(num-1), F=name(num-2))
    R = [("キ 七郎", "茶倶"), ("ク 八郎", "灰中"), ("ケ 九郎", "藍団"),
         ("コ 十郎", "銀中"), ("サ 一一", "銅高"), ("シ 一二", "錫校")]
    for i, (nm, tm) in enumerate(R):
        r = 2 + i * 2
        _set(ws, r, 8, i + 7)         # num at col H(8)
        _set(ws, r, 7, f"({tm})")     # team at G(7)=num-1
        _set(ws, r, 6, nm)            # name at F(6)=num-2


def _build_doubles_stacked(ws):
    # 左半分 stacked: [num][name][team], ペアは2行。num 縦結合。
    pairs = [
        (("サ 一", "白ク"), ("シ 二", "白ク")),     # 同一所属(縦結合) -> 単一表示
        (("ス 三", "黒中"), ("セ 四", "金高")),     # 別所属 -> 連結
        (("ソ 五", "桜ク"), ("タ 六", "桜ク")),
        (("チ 七", "梅中"), ("ツ 八", "竹高")),
        (("テ 九", "松ク"), ("ト 十", "松ク")),
        (("ナ 一一", "藤中"), ("ニ 一二", "杉高")),
    ]
    r = 2
    for i, (p1, p2) in enumerate(pairs):
        _set(ws, r, 1, i + 1); _set(ws, r + 1, 1, i + 1)
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=1)
        _set(ws, r, 2, p1[0]); _set(ws, r + 1, 2, p2[0])
        if p1[1] == p2[1]:
            _set(ws, r, 3, p1[1])
            ws.merge_cells(start_row=r, start_column=3, end_row=r + 1, end_column=3)
        else:
            _set(ws, r, 3, p1[1]); _set(ws, r + 1, 3, p2[1])
        r += 2


def _build_doubles_adjacent(ws):
    # マスター一覧: [num][name1][name2][team1][team2] 横並び, 連番
    rows = [
        ("タ 一", "チ 二", "AAA", "BBB"),
        ("ツ 三", "テ 四", "CCC", "CCC"),   # 同値でも両表示
        ("ト 五", "ナ 六", "DDD", "EEE"),
        ("ニ 七", "ヌ 八", "FFF", "GGG"),
        ("ネ 九", "ノ 十", "HHH", "III"),
        ("ハ 一一", "ヒ 一二", "JJJ", "KKK"),
    ]
    for i, (n1, n2, t1, t2) in enumerate(rows):
        r = 2 + i
        _set(ws, r, 1, i + 1)
        _set(ws, r, 2, n1); _set(ws, r, 3, n2)
        _set(ws, r, 4, f"({t1})"); _set(ws, r, 5, f"({t2})")


def test_pipeline_synthetic():
    print("[pipeline-synthetic]")
    with tempfile.TemporaryDirectory() as d:
        # singles
        wb = Workbook(); ws = wb.active; ws.title = "S"
        _build_singles(ws)
        p = os.path.join(d, "s.xlsx"); wb.save(p)
        g = load_grid(p, "S"); leaves, meta = extract_leaves(g)
        bynum = {L.number: L for L in leaves}
        eq(len(leaves), 12, "singles leaf count")
        eq(bynum[1].members[0], "ア 太郎", "singles L name")
        eq(T.clean_team(bynum[1].member_teams[0]), "赤クラブ", "singles L team")
        eq(bynum[7].members[0], "キ 七郎", "singles R name (mirrored)")
        eq(T.clean_team(bynum[7].member_teams[0]), "茶倶", "singles R team (mirrored)")

        # doubles stacked
        wb = Workbook(); ws = wb.active; ws.title = "DS"
        _build_doubles_stacked(ws)
        p = os.path.join(d, "ds.xlsx"); wb.save(p)
        g = load_grid(p, "DS"); leaves, _ = extract_leaves(g)
        bynum = {L.number: L for L in leaves}
        eq(bynum[1].members, ["サ 一", "シ 二"], "stacked pair members")
        ev = M.build_event(leaves, "DS")
        recs = {p["seed"]: p for p in ev["players"]}
        eq(M.player_display(recs[1])["team"], "白ク", "stacked same-club single team")
        eq(M.player_display(recs[2])["team"], "黒中 / 金高", "stacked diff-club joined team")

        # doubles adjacent (master)
        wb = Workbook(); ws = wb.active; ws.title = "DA"
        _build_doubles_adjacent(ws)
        p = os.path.join(d, "da.xlsx"); wb.save(p)
        g = load_grid(p, "DA"); leaves, _ = extract_leaves(g)
        bynum = {L.number: L for L in leaves}
        eq(bynum[1].members, ["タ 一", "チ 二"], "adjacent pair members")
        ev = M.build_event(leaves, "DA")
        recs = {p["seed"]: p for p in ev["players"]}
        eq(M.player_display(recs[1])["team"], "AAA / BBB", "adjacent two teams")
        eq(M.player_display(recs[2])["team"], "CCC / CCC", "adjacent keeps dup team")

        # emit contract shape
        eq(ev["format"], "doubles", "format detect doubles")
        check(recs[1].get("is_doubles") is True, "is_doubles flag set")
        check("partner_name" in recs[1], "partner_name present")


# ──────────────────────────────────────────────────────────
# 4. 実データ検証(あれば)
# ──────────────────────────────────────────────────────────
def test_real():
    print("[real-data]")
    import json
    F = os.path.expanduser("~/Desktop/卓球関連/ver3.2026 VICTAS杯トーナメント.xlsx")
    J = os.path.expanduser("~/Desktop/卓球関連/json-export/")
    if not os.path.exists(F):
        print("  SKIP (local VICTAS file not found)")
        return
    from collections import Counter
    out = api.parse_workbook(F)
    tot_n = tot_t = tot = 0
    for ev in out["events"]:
        gtf = os.path.join(J, f"victas2026_{ev['event']}.json")
        if not os.path.exists(gtf):
            continue
        gt = json.load(open(gtf, encoding="utf-8"))
        gtl = [(T.normalize_name(p["name"]), T.normalize_name(p.get("team", ""))) for p in gt["players"]]
        myl = []
        for rec in ev["players"]:
            dsp = M.player_display(rec)
            myl.append((T.normalize_name(dsp["name"]), T.normalize_name(dsp["team"])))
        n = sum((Counter(a for a, _ in gtl) & Counter(a for a, _ in myl)).values())
        t = sum((Counter(gtl) & Counter(myl)).values())
        tot_n += n; tot_t += t; tot += len(gtl)
        check(n == len(gtl), f"{ev['event']} names {n}/{len(gtl)}")
        check(t == len(gtl), f"{ev['event']} teams {t}/{len(gtl)}")
    print(f"  real-data totals: NAME {tot_n}/{tot}  TEAM {tot_t}/{tot}")


def main(argv=None):
    argv = argv if argv is not None else sys.argv[1:]
    test_tokens()
    test_grid()
    test_pipeline_synthetic()
    if "--real" in argv:
        test_real()
    print(f"\n{'='*40}\n  PASS {_PASS}  FAIL {_FAIL}\n{'='*40}")
    return 1 if _FAIL else 0


if __name__ == "__main__":
    sys.exit(main())
