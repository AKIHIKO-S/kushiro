#!/usr/bin/env python3
"""
JTTA Excel トーナメント表パーサー v2
=========================================
卓球協会形式のブラケット Excel から、選手リストと対戦ツリーを抽出。

主要アルゴリズム:
  ① 列分類 (Column Classification)
       各列の内容種類を統計的に判定: name / team_paren / team_text / region / seed / round_header
  ② フォーマット検出 (Format Detection)
       master_roster あり? singles? doubles? いずれかを推定
  ③ Entry 抽出 (Entry Extraction)
       name 列のみから選手名を取得。隣接列から team/region/seed を関連付け
  ④ ペア結合 (Doubles Pairing)
       master_roster 優先 (シングル roster あれば最終ソース)
       なければ bracket layout から: 同行隣接列 or 同列連続行 を1ペアに
  ⑤ 罫線追跡 (Border Tracing) -- option
       各 name の右側水平罫線の終端列を計算 → 対戦相手を確定
       「シード選手や予想外の位置」のブラケットを正確に保存

CLI:
    python3 parse_jtta_excel.py FILE.xlsx [--sheet NAME | --all-sheets]
                                          [--output OUT.json]
                                          [--mode seed-list | bracket-tree]
                                          [-v / --verbose]
"""
from __future__ import annotations
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import json
import sys
import re
import argparse
from collections import defaultdict, Counter
from pathlib import Path
from typing import Any


# ─────────────────────────────────────────────────────
# A. テキスト分類
# ─────────────────────────────────────────────────────

# 既知の地域名 (region) - 北海道大会用に十勝/釧路など
KNOWN_REGIONS = set([
    "釧路", "十勝", "北見", "札幌", "千歳", "苫小牧", "根室", "斜里", "名寄",
    "旭川", "函館", "帯広", "石狩", "美幌", "中標津",
    "釧路支部", "札幌支部", "十勝支部", "根室支部",
    "東京", "大阪", "京都", "横浜", "名古屋", "福岡", "仙台",
])

KNOWN_HEADERS = set([
    "決勝", "準決勝", "準々決勝", "予選", "予選リーグ", "本戦", "リーグ戦",
    "男子シングルス", "女子シングルス", "男子ダブルス", "女子ダブルス",
    "混合ダブルス", "団体", "団体戦",
])

# 申込書/様式の固定ラベル — 名前として扱わない
KNOWN_LABELS = set([
    "氏名", "氏名1", "氏名2", "氏名3", "氏名4", "氏名5", "氏名6",
    "名前", "選手名", "代表者", "代表者氏名", "申込責任者", "申込者", "監督", "コーチ",
    "ヘッドコーチ", "顧問", "引率顧問", "引率顧問等", "引率者",
    "出場チーム", "出場選手", "団体名", "学校名", "所属", "所属団体",
    "金額", "参加料", "参加費", "種目", "種別", "区分", "申込区分",
    "メンバー", "メンバー1", "メンバー2", "メンバー3", "メンバー4", "メンバー5",
    "選手1", "選手2", "選手3", "選手4", "選手5",
    "ペア1", "ペア2", "シングル", "ダブル",
    "性別", "男子", "女子", "男", "女",
    "備考", "連絡先", "電話", "電話番号", "メール", "Email", "メールアドレス",
    "印", "申込日", "受付", "受付番号", "通し番号", "番号", "No", "No.",
    "申込", "申し込み", "受付け", "確認", "押印", "記入",
    "一般", "高校", "中学", "中学生", "高校生", "小学", "小学生", "シニア", "ジュニア",
    "計", "合計", "総計", "小計", "総合計",
])

# プレフィックス: ※ で始まる注意書きや指示文
INSTRUCTION_PREFIXES = ("※", "*", "★", "・", "(注)", "（注）", "■", "□", "◯", "●")

TEAM_SUFFIXES = re.compile(
    r'(大学|高校|高等学校|中学校?|小学校?|クラブ|協会|アリーナ|体育館|'
    r'スタジオ|TTC|プラザ|スポーツ|店|社|塾|教室|チーム|チーム名|ジム|'
    r'市役所|町役場|区役所|村役場|庁舎?|商店|会社|株式会社|有限会社|'
    r'病院|医院|クリニック|大会|連盟|連合|館|院|寺|神社|工業|商業|職業|'
    r'銀行|信用|信金|営業所|支店|本店|事業所)$'
)


def classify_value(v: Any) -> str:
    """セル値を分類カテゴリへ"""
    if v is None:
        return "empty"
    s = str(v).strip()
    if not s:
        return "empty"
    # 指示文 (※ で始まる注意書き)
    if any(s.startswith(p) for p in INSTRUCTION_PREFIXES):
        return "instruction"
    # 固定ラベル
    if s in KNOWN_LABELS:
        return "label"
    # 数値
    if re.match(r'^\d+$', s):
        return "int"
    # 数字 (小数や負数も含む)
    if re.match(r'^-?\d+(\.\d+)?$', s):
        return "int"
    # 日付
    if re.match(r'^\d{4}[.\-/]\d+[.\-/]\d+$', s) or re.match(r'^\d+/\d+$', s):
        return "date"
    # 括弧付き = ほぼ確実に team
    if (s.startswith("(") or s.startswith("（")) and \
       (s.endswith(")") or s.endswith("）")):
        inner = s[1:-1].strip()
        # 括弧内が空 or ラベル/指示文 → label
        if not inner or inner in KNOWN_LABELS:
            return "label"
        return "team_paren"
    # ラウンド/セクションヘッダー
    if s in KNOWN_HEADERS:
        return "round_hdr"
    if re.match(r'^ベスト\d+$', s):
        return "round_hdr"
    if re.match(r'^(男子|女子).*ブロック$', s):
        return "section_hdr"
    # 既知地域
    if s in KNOWN_REGIONS:
        return "region"
    # 学校/クラブパターン (括弧なし)
    if TEAM_SUFFIXES.search(s):
        return "team_text"
    # メールアドレス / URL → 名前ではない
    if "@" in s or s.startswith("http"):
        return "other"
    # 電話番号
    if re.match(r'^[\d\-()+ ]+$', s) and re.search(r'\d{3,}', s):
        return "other"
    # 「○○杯」「○○大会」みたいなのは tournament title
    if re.search(r'(杯|大会|戦|の部|選手権|オープン)$', s):
        return "title"
    # 年齢区分・学年 (例: 高校2年, 20歳代以下, 30歳代, シニア)
    if re.match(r'^(\d+歳代?(以上|以下|前後)?|シニア|ジュニア|.{1,5}\d年生?|.{1,5}\d+年|高校\d+年|中学\d+年|小学\d+年|未就学|社会人)$', s):
        return "age_group"
    # 名前らしい (日本語あり、長さ妥当)
    if re.search(r'[ぁ-んァ-ヶー一-龯]', s) and 2 <= len(s) <= 20:
        # ラベル的な末尾 (氏名、選手 など)
        if re.search(r'(氏名|名前|選手|学年|年齢|住所|電話|メール|入力欄?|記入欄?)$', s):
            return "label"
        # 年齢/学年系 (○年生、○代)
        if re.search(r'(年生|歳代?(以上|以下)?|世代|学年)$', s):
            return "age_group"
        return "name"
    # 「a-z 2-3文字」or 「A-Z 2-3文字」(BYEなど) は除外
    if re.match(r'^[A-Za-z]{2,4}$', s) and s.upper() in ("BYE", "TBD", "TBA"):
        return "bye"
    return "other"


def normalize_full_width(s: str) -> str:
    """全角英数字を半角に。 fullname matching に必要"""
    if not s:
        return s
    return s.translate(str.maketrans(
        "０１２３４５６７８９ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ"
        "ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ",
        "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        "abcdefghijklmnopqrstuvwxyz"
    ))


def strip_parens(s: str) -> str:
    s = str(s).strip()
    m = re.match(r'^[(（](.*)[)）]$', s)
    return m.group(1) if m else s


def normalize_name(s: str) -> str:
    """全角空白・連続空白を半角単一空白に正規化 + 全角英数字→半角"""
    if not s:
        return ""
    s = str(s).strip()
    # 全角英数字→半角 (自動修正)
    s = s.translate(str.maketrans(
        "０１２３４５６７８９ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ"
        "ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ",
        "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        "abcdefghijklmnopqrstuvwxyz"
    ))
    # 全角空白を半角に
    s = s.replace("　", " ")
    # 連続空白を 1 つに
    s = re.sub(r'\s+', " ", s)
    # 名前末尾の不要記号 (君、さん、ちゃん、選手) を除去
    s = re.sub(r'(君|くん|さん|ちゃん|選手|くん様|様)$', "", s).strip()
    return s


# ─────────────────────────────────────────────────────
# B. 罫線判定
# ─────────────────────────────────────────────────────

def has_border(cell, side: str) -> bool:
    if cell is None or cell.border is None:
        return False
    b = getattr(cell.border, side, None)
    if b is None or b.style is None:
        return False
    return b.style not in ("none", None)


# ─────────────────────────────────────────────────────
# C. 列分類フェーズ
# ─────────────────────────────────────────────────────

def classify_columns(ws: Worksheet) -> dict[int, dict]:
    """
    全列について、内容タイプの分布と支配的カテゴリを判定。
    戻り値: { col: { 'dominant': 'name', 'stats': {...}, 'total': N, 'name_count': N } }
    """
    col_stats: dict[int, Counter] = defaultdict(Counter)
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            t = classify_value(cell.value)
            if t != "empty":
                col_stats[cell.column][t] += 1

    result = {}
    for col, stats in col_stats.items():
        total = sum(stats.values())
        if total < 2:
            continue
        sorted_kinds = stats.most_common()
        dominant = sorted_kinds[0][0]
        name_count = stats.get("name", 0)
        # name が混じってる team_text 列 → team_text の方が信頼度高い
        if dominant == "name" and stats.get("team_text", 0) > total * 0.4:
            dominant = "team_text"
        # label が大半なら label 列 (申込書テンプレート)
        if dominant == "label" or stats.get("label", 0) > total * 0.5:
            dominant = "label"
        # instruction が大半なら instruction 列 (注意書き)
        if stats.get("instruction", 0) > total * 0.5:
            dominant = "instruction"
        # 名前候補が少なすぎる (< 4) → name 扱いしない (申込書テンプレートの数行サンプルを除外)
        if dominant == "name" and name_count < 4:
            dominant = "other"
        result[col] = {
            "dominant": dominant,
            "stats": dict(stats),
            "total": total,
            "name_count": name_count,
        }
    return result


def detect_surname_givenname_split(ws, col_class, name_cols):
    """
    隣接する 2 つの name 列が「苗字」+「名前」のように分割されているか判定。
    判定基準:
      - 同じ行に両方の値がある率が 70% 以上
      - 各セルの平均文字数が短い (3 字以下)
    返り値: [(surname_col, given_col), ...] のリスト
    """
    pairs = []
    cols = sorted(name_cols)
    used = set()
    for i, c1 in enumerate(cols):
        if c1 in used:
            continue
        if i + 1 >= len(cols):
            break
        c2 = cols[i + 1]
        # 隣接 (距離 == 1) のみ
        if c2 - c1 != 1:
            continue

        same_row_count = 0
        c1_count = 0
        c2_count = 0
        c1_avg_len = 0
        c2_avg_len = 0
        for r in range(1, ws.max_row + 1):
            v1 = ws.cell(row=r, column=c1).value
            v2 = ws.cell(row=r, column=c2).value
            if v1 and classify_value(v1) == "name":
                c1_count += 1
                c1_avg_len += len(str(v1).strip())
            if v2 and classify_value(v2) == "name":
                c2_count += 1
                c2_avg_len += len(str(v2).strip())
            if v1 and v2 and classify_value(v1) == "name" and classify_value(v2) == "name":
                same_row_count += 1

        if c1_count == 0 or c2_count == 0:
            continue
        c1_avg_len /= max(c1_count, 1)
        c2_avg_len /= max(c2_count, 1)
        # 同じ行に両方 (70%以上) かつ 平均 3 文字以下 → 苗字/名前 split
        threshold_ratio = same_row_count / max(c1_count, c2_count, 1)
        if threshold_ratio >= 0.7 and c1_avg_len <= 3 and c2_avg_len <= 3:
            pairs.append((c1, c2))
            used.add(c1)
            used.add(c2)
    return pairs


def merge_surname_givenname(ws, col_class, pairs):
    """
    苗字/名前ペアを「name」列扱いに統合 (内部マージ用テーブル返却)。
    戻り値: { surname_col: { 'paired_col': given_col, 'merged': True } }
    """
    mapping = {}
    for s, g in pairs:
        mapping[s] = {"paired_col": g, "merged": True}
    return mapping


# ─────────────────────────────────────────────────────
# D. 形式判定 & エントリー抽出
# ─────────────────────────────────────────────────────

def detect_blocks(name_cols: list, is_doubles: bool, master_cols: dict | None = None) -> list:
    """
    name 列をブロックにグループ化する。
    JTTA形式: 1ブロックは左右2本の name 列 (近接列で対) を含む。
    例 (男子S 4ブロック): [(2,24)='A', (28,50)='B', (54,76)='C', (80,102)='D']
    master_roster_cols は除外。

    返り値: [{'block': 'A', 'cols': [c1, c2]}, ...]
    """
    cols = sorted(name_cols)
    # master_roster の列は除外
    if master_cols:
        excluded = set(master_cols.values())
        cols = [c for c in cols if c not in excluded]

    blocks = []
    used = set()
    block_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

    if is_doubles:
        # ダブルスは1ペア (col1, col2) が連続列
        # ブロック内では (col1, col2) と (col3, col4) のような近接ペア構造
        # cols は隣接ペアのリストとして扱う
        # 例: [5,6, 26,27, 34,35, 55,56] → A=(5-6), B=(26-27)... ではなく
        #     [(5,6)] + [(26,27)] + [(34,35)] + [(55,56)] が各ブロック
        i = 0
        bi = 0
        while i < len(cols):
            if i + 1 < len(cols) and cols[i+1] - cols[i] == 1:
                # 同行ペア (col, col+1) は1グループ
                blocks.append({'block': block_letters[bi] if bi < len(block_letters) else f"B{bi+1}",
                              'cols': [cols[i], cols[i+1]]})
                used.add(cols[i]); used.add(cols[i+1])
                i += 2
                bi += 1
            else:
                i += 1
        return blocks

    # シングルス: ブロック内に左右の name 列 (距離 < 30) があるとペア化
    i = 0
    bi = 0
    while i < len(cols):
        c = cols[i]
        if c in used:
            i += 1; continue
        # 同じブロックの「右側」列を探す: 距離 5-30 程度
        partner = None
        for j in range(i+1, len(cols)):
            if cols[j] in used: continue
            d = cols[j] - c
            if 5 <= d <= 30:
                partner = cols[j]
                break
            if d > 30:
                break
        block = block_letters[bi] if bi < len(block_letters) else f"B{bi+1}"
        if partner:
            blocks.append({'block': block, 'cols': [c, partner]})
            used.add(c); used.add(partner)
        else:
            blocks.append({'block': block, 'cols': [c]})
            used.add(c)
        bi += 1
        i += 1
    return blocks


def _col_to_block(blocks: list, col: int) -> str:
    for b in blocks:
        if col in b['cols']:
            return b['block']
    return ""


def detect_format(col_class: dict[int, dict], sheet_name: str, ws=None) -> dict:
    """
    シート全体の形式を判定。
    戻り値:
      {
        'is_doubles': bool,
        'name_cols': [n1, n2, ...],
        'paired_name_cols': [(n1a, n1b), (n2a, n2b), ...],  # ダブルスの隣接列ペア
        'master_roster_cols': dict or None,  # マスター roster がある場合
        'surname_given_pairs': [(s,g), ...],  # 苗字+名前 split ペア
      }
    """
    name_cols = [c for c, info in col_class.items() if info["dominant"] == "name"]
    name_cols.sort()

    # シート名から「ダブルス」判定; 「団体」「ダブル」「ペア」を含む
    is_doubles_sheet = any(k in sheet_name for k in ["ダブルス", "団体", "ダブル", "ペア", "Doubles", "Team"])

    # 苗字+名前 split 列ペアを検出 (シングルス向け)
    surname_given_pairs = []
    if ws is not None and not is_doubles_sheet:
        surname_given_pairs = detect_surname_givenname_split(ws, col_class, name_cols)

    # surname-given にマージされた列を除外して paired_name_cols 検出
    sg_cols = set([c for pair in surname_given_pairs for c in pair])

    # 隣接列の name ペアを検出 (ダブルス形式)
    paired_name_cols = []
    used = set(sg_cols)
    for c in name_cols:
        if c in used:
            continue
        if (c + 1) in name_cols and (c + 1) not in used:
            paired_name_cols.append((c, c + 1))
            used.add(c)
            used.add(c + 1)

    # master roster 候補
    master_roster_cols = None
    if is_doubles_sheet and paired_name_cols:
        # ダブルス: 隣接 name ペアで件数が他の2倍以上多いものが master roster
        pair_counts = []
        for (a, b) in paired_name_cols:
            cnt = min(col_class[a]["total"], col_class[b]["total"])
            pair_counts.append(((a, b), cnt))
        pair_counts.sort(key=lambda x: -x[1])
        if len(pair_counts) >= 2 and pair_counts[0][1] >= pair_counts[1][1] * 1.5:
            master_roster_cols = {
                "name1_col": pair_counts[0][0][0],
                "name2_col": pair_counts[0][0][1],
            }
    elif not is_doubles_sheet and name_cols:
        # シングルス: 単独 name 列で件数が他の2倍以上ある場合 master roster
        # ただし surname-given ペアが検出されている場合はそれを優先
        if not surname_given_pairs:
            col_counts = [(c, col_class[c]["total"]) for c in name_cols if c not in used]
            col_counts.sort(key=lambda x: -x[1])
            if len(col_counts) >= 2 and col_counts[0][1] >= col_counts[1][1] * 2:
                master_roster_cols = {"name_col": col_counts[0][0]}

    return {
        "is_doubles": is_doubles_sheet,
        "name_cols": name_cols,
        "paired_name_cols": paired_name_cols,
        "master_roster_cols": master_roster_cols,
        "surname_given_pairs": surname_given_pairs,
    }


def extract_entries(ws: Worksheet, col_class: dict, fmt: dict, verbose: bool = False) -> list[dict]:
    """
    各 name セルを取り出し、隣接 cell から team/region/seed を補完。
    重複除外も実施。
    """
    is_doubles = fmt["is_doubles"]
    master = fmt["master_roster_cols"]
    sg_pairs = fmt.get("surname_given_pairs", [])

    # ──── ダブルス + master roster あり: master だけ使う
    if is_doubles and master:
        if verbose:
            print(f"  [extract] using master roster cols={master}", file=sys.stderr)
        return _extract_doubles_from_roster(ws, master, col_class)

    # ──── ダブルス + 列ペア あり: 各ペアを使う
    if is_doubles and fmt["paired_name_cols"]:
        return _extract_doubles_from_pairs(ws, fmt["paired_name_cols"], col_class, verbose)

    # ──── ダブルス + 列ペアなし + master なし: 同列連続行で paring
    if is_doubles:
        if verbose:
            print("  [extract] doubles fallback: column-stack pairing", file=sys.stderr)
        return _extract_doubles_from_column_stack(ws, fmt["name_cols"], col_class)

    # ──── シングルス (苗字+名前 split があれば結合)
    return _extract_singles(ws, fmt["name_cols"], col_class, verbose, sg_pairs)


def _extract_doubles_from_column_stack(ws, name_cols, col_class):
    """同列で連続行 (row N, row N+1 or N+2) を 1ペア化 (女子ダブルス形式)。"""
    entries = []
    seen = set()
    for nc in name_cols:
        team_col = _find_adjacent(name_cols, col_class, nc, ["team_paren", "team_text"])
        seed_col = _find_adjacent(name_cols, col_class, nc, ["int"])

        rows_with_names = []
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=nc).value
            if v and classify_value(v) == "name":
                rows_with_names.append(r)

        i = 0
        while i < len(rows_with_names):
            r1 = rows_with_names[i]
            # 次の name 行が 1-2 行差ならペア化
            if i + 1 < len(rows_with_names) and rows_with_names[i + 1] - r1 <= 2:
                r2 = rows_with_names[i + 1]
                name1 = normalize_name(ws.cell(row=r1, column=nc).value)
                name2 = normalize_name(ws.cell(row=r2, column=nc).value)
                pair_key = tuple(sorted([name1, name2]))
                if pair_key not in seen:
                    seen.add(pair_key)
                    team1 = ""
                    team2 = ""
                    if team_col:
                        t1v = ws.cell(row=r1, column=team_col).value
                        if t1v: team1 = strip_parens(str(t1v).strip())
                        t2v = ws.cell(row=r2, column=team_col).value
                        if t2v: team2 = strip_parens(str(t2v).strip())
                    seed = 0
                    if seed_col:
                        sv = ws.cell(row=r1, column=seed_col).value
                        try:
                            n = int(sv) if sv is not None else 0
                            if 1 <= n <= 999:
                                seed = n
                        except (TypeError, ValueError):
                            pass
                    entries.append({
                        "name": name1, "team": team1,
                        "partner_name": name2, "partner_team": team2,
                        "seed": seed, "is_doubles": True,
                        "_row": r1, "_col": nc,
                    })
                i += 2
            else:
                i += 1
    return entries


def _extract_singles(ws, name_cols, col_class, verbose, sg_pairs=None):
    """各 name 列からシングルス選手を取得 (sg_pairs: [(surname_col, given_col), ...])"""
    entries = []
    seen = set()  # (name, team) で dedup

    # 苗字+名前 split がある列ペアは特別に扱う
    sg_pairs = sg_pairs or []
    sg_lookup = {s: g for s, g in sg_pairs}
    paired_cols = set([c for pair in sg_pairs for c in pair])

    # まず通常の name 列処理 (paired から「苗字」列のみ処理、「名前」列はスキップ)
    given_only_cols = set(g for _, g in sg_pairs)

    for nc in name_cols:
        # 名前のみ列は別ペアから処理されるのでスキップ
        if nc in given_only_cols:
            continue
        # 苗字列の場合、隣接「名前」列と結合
        merged_with = sg_lookup.get(nc)

        # 隣接列で team/seed を推測 (paired_cols は除外して探索)
        search_cols = name_cols if not merged_with else [c for c in name_cols if c != merged_with]
        team_col = _find_adjacent(search_cols, col_class, nc, ["team_paren", "team_text"])
        seed_col = _find_adjacent(search_cols, col_class, nc, ["int"])
        region_col = _find_adjacent(search_cols, col_class, nc, ["region"])

        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=nc).value
            if not v: continue
            t = classify_value(v)
            if t != "name": continue
            surname = normalize_name(v)

            # 苗字+名前 split の場合 → 結合
            if merged_with:
                gv = ws.cell(row=r, column=merged_with).value
                if gv and classify_value(gv) == "name":
                    given = normalize_name(gv)
                    name = f"{surname} {given}"
                else:
                    # 名前が無い行は不完全な行 → スキップ (ヘッダー等)
                    continue
            else:
                name = surname

            team = ""
            if team_col:
                tv = ws.cell(row=r, column=team_col).value
                if tv:
                    team = strip_parens(str(tv).strip())

            region = ""
            if region_col:
                rv = ws.cell(row=r, column=region_col).value
                if rv: region = str(rv).strip()

            seed = 0
            if seed_col:
                sv = ws.cell(row=r, column=seed_col).value
                try:
                    n = int(sv) if sv is not None else 0
                    if 1 <= n <= 999:
                        seed = n
                except (TypeError, ValueError): pass

            key = (name, team)
            if key in seen:
                continue
            seen.add(key)
            entries.append({
                "name": name, "team": team, "region": region, "seed": seed,
                "_row": r, "_col": nc,
            })
    return entries


def _extract_doubles_from_roster(ws, master, col_class):
    """master roster から ダブルスペアを取得 (最も信頼度高い source)"""
    n1c = master["name1_col"]
    n2c = master["name2_col"]
    seed_col = _find_adjacent([n1c, n2c], col_class, n1c, ["int"])
    team1_col = _find_adjacent([n1c, n2c], col_class, n2c, ["team_paren"])
    team2_col = _find_adjacent([n1c, n2c], col_class, team1_col or n2c, ["team_paren"]) if team1_col else None
    region_col = _find_adjacent([n1c, n2c, team1_col or 0, team2_col or 0], col_class,
                                 team2_col or n2c, ["region"])

    entries = []
    seen = set()
    for r in range(1, ws.max_row + 1):
        v1 = ws.cell(row=r, column=n1c).value
        v2 = ws.cell(row=r, column=n2c).value
        if not v1 or not v2: continue
        if classify_value(v1) != "name" or classify_value(v2) != "name": continue

        name1 = normalize_name(v1)
        name2 = normalize_name(v2)
        if (name1, name2) in seen or (name2, name1) in seen:
            continue
        seen.add((name1, name2))

        team1 = strip_parens(str(ws.cell(row=r, column=team1_col).value).strip()) \
                if team1_col and ws.cell(row=r, column=team1_col).value else ""
        team2 = strip_parens(str(ws.cell(row=r, column=team2_col).value).strip()) \
                if team2_col and ws.cell(row=r, column=team2_col).value else ""
        region = str(ws.cell(row=r, column=region_col).value).strip() \
                 if region_col and ws.cell(row=r, column=region_col).value else ""
        seed = 0
        if seed_col:
            sv = ws.cell(row=r, column=seed_col).value
            try:
                n = int(sv) if sv is not None else 0
                if 1 <= n <= 999: seed = n
            except (TypeError, ValueError): pass

        entries.append({
            "name": name1, "team": team1,
            "partner_name": name2, "partner_team": team2,
            "region": region, "seed": seed,
            "is_doubles": True,
            "_row": r, "_col": n1c,
        })
    return entries


def _extract_doubles_from_pairs(ws, pairs, col_class, verbose):
    """bracket layout から ダブルスペアを取得 (master roster がない場合)"""
    entries = []
    seen = set()
    for (n1c, n2c) in pairs:
        team_col = _find_adjacent([n1c, n2c], col_class, n2c, ["team_paren", "team_text"])
        team2_col = _find_adjacent([n1c, n2c, team_col or 0], col_class,
                                    team_col or n2c, ["team_paren", "team_text"]) if team_col else None
        seed_col = _find_adjacent([n1c, n2c], col_class, n1c, ["int"])

        for r in range(1, ws.max_row + 1):
            v1 = ws.cell(row=r, column=n1c).value
            v2 = ws.cell(row=r, column=n2c).value
            if not v1 and not v2:
                continue
            # 片方だけ name のときも記録 (パートナー欠落として警告)
            c1_kind = classify_value(v1) if v1 else "empty"
            c2_kind = classify_value(v2) if v2 else "empty"
            if c1_kind != "name" and c2_kind != "name":
                continue

            name1 = normalize_name(v1) if c1_kind == "name" else ""
            name2 = normalize_name(v2) if c2_kind == "name" else ""
            if (name1, name2) in seen or (name2, name1) in seen:
                continue
            seen.add((name1, name2))

            team1 = ""
            team2 = ""
            if team_col:
                tv = ws.cell(row=r, column=team_col).value
                if tv: team1 = strip_parens(str(tv).strip())
            if team2_col:
                tv2 = ws.cell(row=r, column=team2_col).value
                if tv2: team2 = strip_parens(str(tv2).strip())

            seed = 0
            if seed_col:
                sv = ws.cell(row=r, column=seed_col).value
                try:
                    n = int(sv) if sv is not None else 0
                    if 1 <= n <= 999: seed = n
                except (TypeError, ValueError): pass

            entries.append({
                "name": name1 or name2, "team": team1 or team2,
                "partner_name": name2 if name1 else "",
                "partner_team": team2 if name1 else "",
                "seed": seed, "is_doubles": True,
                "_row": r, "_col": n1c,
                "_pair_complete": bool(name1 and name2),
            })
    return entries


def _find_adjacent(name_cols, col_class, anchor_col, target_kinds, max_dist=5):
    """anchor_col から近い距離で指定 dominant kind の列を探す"""
    candidates = []
    for col, info in col_class.items():
        if info["dominant"] in target_kinds:
            d = abs(col - anchor_col)
            if d <= max_dist:
                candidates.append((d, col))
    candidates.sort()
    return candidates[0][1] if candidates else None


# ─────────────────────────────────────────────────────
# E. 罫線追跡フェーズ (bracket-tree mode)
# ─────────────────────────────────────────────────────

def trace_match_pairings(ws, entries, max_step=120):
    """
    各 entry の name セルから右側に水平罫線を追跡し、
    同じ終端列に達する2つの entry を round 1 対戦ペアとする。
    シード位置や予想外配置に対応するため、これにより actual bracket layout を保存。

    戻り値: [{ 'p1_idx': i, 'p2_idx': j, 'endpoint_col': c, 'mid_row': r }]
    """
    # 各 entry の水平線終端を計算
    endpoints = []
    for idx, e in enumerate(entries):
        r = e["_row"]
        c = e["_col"]
        # name cell の下端罫線を右に追跡
        end_col = c
        for step in range(max_step):
            cur_cell = ws.cell(row=r, column=end_col)
            next_cell = ws.cell(row=r, column=end_col + 1)
            below = ws.cell(row=r + 1, column=end_col)
            below_next = ws.cell(row=r + 1, column=end_col + 1)
            # 水平線継続条件 (bottom of this row OR top of next row)
            cont = (has_border(cur_cell, "bottom") or has_border(below, "top")
                    or has_border(next_cell, "bottom") or has_border(below_next, "top"))
            if not cont:
                break
            end_col += 1
        endpoints.append((idx, r, end_col))

    # 同じ終端列に達する entry を集約 → 対戦ペア
    by_endpoint = defaultdict(list)
    for idx, r, ec in endpoints:
        by_endpoint[ec].append((r, idx))

    pairs = []
    for ec, items in by_endpoint.items():
        items.sort()
        # 隣接する2件を round 1 ペアとする (3件以上は近接行で集約)
        for i in range(0, len(items) - 1, 2):
            r1, i1 = items[i]
            r2, i2 = items[i + 1]
            if r2 - r1 <= 6:  # 妥当な行間内のみ
                pairs.append({
                    "p1_idx": i1, "p2_idx": i2,
                    "endpoint_col": ec,
                    "mid_row": (r1 + r2) // 2,
                })
    return pairs


def build_bracket_tree(entries, pairings, total_rounds):
    """
    pairings (round-1 matches) を元に、トーナメントツリーを構築。
    各ラウンドの match を bracket_pos でインデックス化。
    """
    matches_by_round = [[] for _ in range(total_rounds)]
    # round 1 を pairings から
    pairings_sorted = sorted(pairings, key=lambda p: p["mid_row"])
    for pos, p in enumerate(pairings_sorted):
        e1 = entries[p["p1_idx"]]
        e2 = entries[p["p2_idx"]]
        matches_by_round[0].append({
            "bracket_round": 1,
            "bracket_pos": pos,
            "player1_name": e1["name"],
            "player1_team": e1.get("team", ""),
            "player2_name": e2["name"],
            "player2_team": e2.get("team", ""),
        })
        if e1.get("is_doubles"):
            matches_by_round[0][-1]["player1_partner_name"] = e1.get("partner_name", "")
            matches_by_round[0][-1]["player2_partner_name"] = e2.get("partner_name", "")

    # 上位ラウンドは空 (アプリ側で next_match を自動リンク)
    for r in range(1, total_rounds):
        num = max(1, len(matches_by_round[0]) // (2 ** r))
        for pos in range(num):
            matches_by_round[r].append({
                "bracket_round": r + 1,
                "bracket_pos": pos,
                "player1_name": "",
                "player2_name": "",
            })

    flat = []
    for rnd in matches_by_round:
        flat.extend(rnd)
    return flat


# ─────────────────────────────────────────────────────
# F. メイン抽出 + 出力
# ─────────────────────────────────────────────────────

def parse_sheet(xlsx_path: Path, sheet_name: str,
                event_name: str | None = None,
                mode: str = "seed-list",
                verbose: bool = False) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"シート '{sheet_name}' が見つかりません")
    ws = wb[sheet_name]
    if verbose:
        print(f"[{sheet_name}] {ws.max_row}行 x {ws.max_column}列", file=sys.stderr)

    # ① 列分類
    col_class = classify_columns(ws)
    name_cols = sorted([c for c, info in col_class.items() if info["dominant"] == "name"])
    if verbose:
        print(f"  name列: {name_cols}", file=sys.stderr)

    # ② 形式判定
    fmt = detect_format(col_class, sheet_name, ws=ws)
    if verbose:
        print(f"  is_doubles={fmt['is_doubles']}, "
              f"paired={fmt['paired_name_cols']}, "
              f"master={fmt['master_roster_cols']}, "
              f"surname_given={fmt.get('surname_given_pairs')}", file=sys.stderr)

    # ②-b ブロック検出
    blocks = detect_blocks(name_cols, fmt["is_doubles"], fmt.get("master_roster_cols"))
    if verbose:
        for b in blocks:
            print(f"  block {b['block']}: cols={b['cols']}", file=sys.stderr)

    # ③ 抽出
    entries = extract_entries(ws, col_class, fmt, verbose)
    # 各 entry に block を付与 (_col から判定)
    for e in entries:
        e["block"] = _col_to_block(blocks, e.get("_col", 0))
    if verbose:
        from collections import Counter
        bcnt = Counter(e.get("block","") for e in entries)
        print(f"  抽出 entries: {len(entries)} 件 (ブロック分布={dict(bcnt)})", file=sys.stderr)

    # ④ seed 順ソート
    entries_sorted = sorted(entries, key=lambda e: (e.get("seed", 0) or 9999, e.get("_row", 0)))

    # ⑤ 出力モード分岐
    if mode == "bracket-tree":
        # 罫線追跡 → 実 bracket layout を保存
        pairings = trace_match_pairings(ws, entries_sorted)
        # 総ラウンド数
        n = len(pairings) * 2 if pairings else len(entries_sorted)
        import math
        total_rounds = max(1, math.ceil(math.log2(max(n, 2)))) if n > 0 else 1
        matches = build_bracket_tree(entries_sorted, pairings, total_rounds)
        if verbose:
            print(f"  trace: {len(pairings)} round1ペア, {len(matches)} 試合", file=sys.stderr)
        return {
            "format": "tabletennis-bracket-v1",
            "event": event_name or sheet_name,
            "regenerate": True,
            "auto_link_to_players": False,
            "matches": matches,
            "_meta": {
                "source": str(xlsx_path), "sheet": sheet_name,
                "mode": "bracket-tree",
                "is_doubles": fmt["is_doubles"],
                "entries_extracted": len(entries_sorted),
                "round1_pairs": len(pairings),
            },
        }
    else:
        # seed-list モード (default)
        players = []
        warnings = []
        auto_fixed = []  # 自動修正された項目
        for e in entries_sorted:
            p = {"name": e["name"]}
            if e.get("team"): p["team"] = e["team"]
            if e.get("seed"): p["seed"] = e["seed"]
            if e.get("block"): p["block"] = e["block"]
            if e.get("region"): p["region"] = e["region"]
            if e.get("is_doubles"):
                p["is_doubles"] = True
                if e.get("partner_name"): p["partner_name"] = e["partner_name"]
                if e.get("partner_team"): p["partner_team"] = e["partner_team"]
                # ダブルスでパートナー無しは警告
                if not e.get("partner_name"):
                    warnings.append({
                        "type": "missing_partner",
                        "player": e["name"],
                        "row": e.get("_row"),
                        "message": f"ダブルス選手「{e['name']}」のパートナーが見つかりません。Excel の同行・隣接列を確認してください。",
                        "severity": "error",
                    })
            # フルネーム検出: 名前にスペースを含まない (苗字のみ)
            elif " " not in e["name"] and not fmt["is_doubles"]:
                warnings.append({
                    "type": "surname_only",
                    "player": e["name"],
                    "row": e.get("_row"),
                    "message": f"「{e['name']}」は苗字のみの可能性があります。フルネームが別列にある場合は手動で修正してください。",
                    "severity": "warn",
                })
            players.append(p)
        # 全体警告
        if len(players) < 4 and len(players) > 0:
            warnings.append({
                "type": "too_few_entries",
                "message": f"抽出選手数が {len(players)} 名と少ないです。記載例シートやサンプルデータの可能性があります。シート名を確認してください。",
                "severity": "warn",
            })
        if len(players) == 0:
            warnings.append({
                "type": "no_entries",
                "message": "選手が 1 名も検出されませんでした。シート名が間違っているか、Excel のレイアウトが想定と異なります。",
                "severity": "error",
            })
        # 苗字+名前が分かれていた場合の自動修正通知
        if fmt.get("surname_given_pairs"):
            auto_fixed.append({
                "type": "merged_surname_given",
                "message": f"苗字と名前が別列に分かれていたので自動結合しました ({len(fmt['surname_given_pairs'])} ペア)",
                "cols": fmt["surname_given_pairs"],
            })
        return {
            "format": "tabletennis-seed-list-v1",
            "event": event_name or sheet_name,
            "regenerate": True,
            "auto_link_to_players": False,
            "players": players,
            "warnings": warnings,
            "auto_fixed": auto_fixed,
            "_meta": {
                "source": str(xlsx_path), "sheet": sheet_name,
                "mode": "seed-list",
                "is_doubles": fmt["is_doubles"],
                "name_cols_detected": name_cols,
                "master_roster": fmt.get("master_roster_cols"),
                "surname_given_pairs": fmt.get("surname_given_pairs"),
                "warning_count": len(warnings),
                "auto_fixed_count": len(auto_fixed),
            },
        }


# ─────────────────────────────────────────────────────
# G. CLI
# ─────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="JTTA形式 Excel → tabletennis JSON 変換 (高精度パーサー)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
モード:
  seed-list      (default) 選手リスト形式。アプリ側で標準シーディングで bracket 生成。
  bracket-tree   罫線追跡で実 bracket レイアウトを保存。シード位置・予想外位置に対応。
""")
    p.add_argument("xlsx", help="入力 .xlsx ファイル")
    p.add_argument("--sheet", help="シート名 (省略時は全シート出力)")
    p.add_argument("--event-name", help="イベント名 (省略時はシート名)")
    p.add_argument("--output", help="出力ファイル")
    p.add_argument("--all-sheets", action="store_true", help="全シートを一括")
    p.add_argument("--mode", choices=["seed-list", "bracket-tree"], default="seed-list")
    p.add_argument("--skip-sheets", default="Sheet1,Sheet,審判員",
                   help="無視シート (カンマ区切り)")
    p.add_argument("-v", "--verbose", action="store_true")
    args = p.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"[ERROR] {xlsx_path} が見つかりません", file=sys.stderr)
        sys.exit(1)

    skip = set(args.skip_sheets.split(","))

    if args.all_sheets:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        brackets = []
        for sn in wb.sheetnames:
            if sn in skip: continue
            try:
                b = parse_sheet(xlsx_path, sn, event_name=sn,
                                mode=args.mode, verbose=args.verbose)
                brackets.append(b)
                key = "players" if args.mode == "seed-list" else "matches"
                cnt = len(b.get(key, []))
                if args.verbose:
                    print(f"  {sn}: {cnt} 件", file=sys.stderr)
            except Exception as e:
                print(f"[WARN] {sn}: {e}", file=sys.stderr)
        out = {
            "format": "tabletennis-tournament-v1",
            "tournament": {"name": xlsx_path.stem},
            "brackets": brackets,
        }
    else:
        sheet = args.sheet or openpyxl.load_workbook(xlsx_path, data_only=True).sheetnames[0]
        out = parse_sheet(xlsx_path, sheet, event_name=args.event_name,
                          mode=args.mode, verbose=args.verbose)

    if args.output:
        Path(args.output).write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[OK] 書き出し: {args.output}", file=sys.stderr)
    else:
        print(json.dumps(out, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
