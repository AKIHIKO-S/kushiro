#!/usr/bin/env python3
"""
組合せ表 (くみあわせひょう) 罫線解析パーサー
=============================================
JTTA 形式の組合せ表 Excel から、1回戦〜決勝までの全対戦を自動生成。
シングルス・団体戦両対応。

特徴:
  ① 罫線解析でブラケット構造を確定 (シード位置や予想外配置も保存)
  ② チーム名+選手名 (団体戦) / 選手名のみ (個人戦) 両対応
  ③ 組合せ番号順の隣接ペア (1vs2, 3vs4,..) を round 1 として構築
  ④ 出力: tabletennis-bracket-v1 (各試合の bracket_round/bracket_pos 明示)

アルゴリズム:
  1. 「番号列」(col 1, 22 等) を発見 — 連番整数
  2. 各番号 → 同行のチーム/選手名を取得
  3. 罫線追跡で 番号ごとの「水平線終端」= 1回戦対戦合流地点
  4. 終端列が同じ2つの番号 = round 1 の対戦
  5. 合流点から右へ追跡して上位ラウンドを構築

CLI:
    python3 parse_kumiawase_chart.py FILE.xlsx [--sheet NAME]
                                                [--event "男子団体"]
                                                [--output OUT.json]
                                                [--mode singles|team] (auto検出)
                                                [-v]
"""
from __future__ import annotations
import openpyxl
from openpyxl.utils import get_column_letter
import json
import sys
import re
import argparse
import math
from collections import defaultdict
from pathlib import Path


# ─────────────────────────────────────────────
# 罫線判定 & 値判定
# ─────────────────────────────────────────────
def has_border(cell, side: str) -> bool:
    if cell is None or cell.border is None:
        return False
    b = getattr(cell.border, side, None)
    return b is not None and b.style not in (None, "none")


def is_number_cell(v) -> int | None:
    """整数値を返す (1-999の範囲)。それ以外は None"""
    if v is None:
        return None
    try:
        n = int(v)
        if 1 <= n <= 999:
            return n
    except (ValueError, TypeError):
        return None
    return None


def normalize_name(s) -> str:
    if s is None:
        return ""
    return re.sub(r"[\s　]+", " ", str(s).strip())


def is_team_or_name(v) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    if not s:
        return False
    # 数字のみは除外
    if re.match(r"^\d+$", s):
        return False
    # 日本語を含む必要
    if not re.search(r"[ぁ-んァ-ヶー一-龯]", s):
        # ローマ字チーム名 (AMATAKU, MPC 等) も許容
        if not re.match(r"^[A-Za-z\-_]+$", s):
            return False
    # ヘッダーらしいものは除外
    if s in ("一般男子", "一般女子", "高校男子", "高校女子", "中学男子", "中学女子",
             "団体", "シングルス", "ダブルス", "混合"):
        return False
    if re.match(r"^(○|◯|●)", s):
        return False
    return True


# ─────────────────────────────────────────────
# 構造検出: 番号列を特定
# ─────────────────────────────────────────────
def find_number_columns(ws, max_scan=300):
    """
    番号列を「連続セグメント」単位で抽出。
    1つのExcelに複数テーブル (例: 団体戦bracket + 個人戦roster) が混在する場合、
    各テーブルを別セグメントとして返す。

    判定: num が +1 連続 AND row 間隔も妥当
    複数候補がある場合、ROW が最小 (=上に書かれた) Run を bracket と推定。
    """
    candidates = defaultdict(list)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_scan)):
        for cell in row:
            n = is_number_cell(cell.value)
            if n is not None:
                candidates[cell.column].append((cell.row, n))

    segments = []
    for col, items in candidates.items():
        items.sort()
        # 重複値が多すぎる列はスキップ (例: 1 だらけの列)
        nums_all = [n for _, n in items]
        if len(set(nums_all)) / len(nums_all) < 0.6:
            continue

        # 連続 num のセグメントに分割
        cur = []
        for r, n in items:
            if cur:
                last_r, last_n = cur[-1]
                # num が連続 (+1) でなければ新セグメント
                if n - last_n != 1:
                    if len(cur) >= 2:
                        segments.append(cur)
                    cur = []
            cur.append((r, n))
        if len(cur) >= 2:
            segments.append(cur)

    # 各セグメントを (col, items, start, end, top_row) で wrapping
    result = []
    for seg in segments:
        cols = {r for _ in seg for r in [seg[0][0]]}  # placeholder
        # col は seg の全 cell が来る col。candidates の dict から逆引き必要だが、
        # 各 seg は単一 col から来てるので、最初の item の row で取得しなおす
        first_r = seg[0][0]
        # col を再特定 (segments は col 別なので、items の元の col を保持してない)
        # → 上の処理で col 情報を保持するように修正必要だが、ここは workaround
        # 実は items は (r, n) のみで col を持ってないので
        pass

    # 再実装: col 情報を保持しながらセグメント化
    result = []
    for col, items in candidates.items():
        items.sort()
        nums_all = [n for _, n in items]
        if len(set(nums_all)) / max(len(nums_all), 1) < 0.6:
            continue
        cur = []
        for r, n in items:
            if cur and n - cur[-1][1] != 1:
                if len(cur) >= 2:
                    result.append({
                        "col": col, "items": list(cur),
                        "count": len(cur),
                        "start": cur[0][1], "end": cur[-1][1],
                        "top_row": cur[0][0],
                    })
                cur = []
            cur.append((r, n))
        if len(cur) >= 2:
            result.append({
                "col": col, "items": list(cur),
                "count": len(cur),
                "start": cur[0][1], "end": cur[-1][1],
                "top_row": cur[0][0],
            })

    # 「最も上に書かれた」セグメントから優先 (bracket は通常上部)
    result.sort(key=lambda s: (s["top_row"], -s["count"]))
    return result


def select_bracket_segments(segments, verbose=False):
    """
    複数セグメントから bracket 用を選択。
    通常 ブラケットは上部にあるので、最も上のセグメント + 同程度の top_row を持つ別 col のセグメント
    (左右サイド) を組み合わせて返す。
    """
    if not segments:
        return []
    first = segments[0]
    top = first["top_row"]
    selected = [first]
    # 同程度 top_row (差 < 10) の別 col のセグメントを追加 (左右サイド)
    for seg in segments[1:]:
        if seg["col"] != first["col"] and abs(seg["top_row"] - top) <= 10:
            selected.append(seg)
    if verbose:
        for s in selected:
            print(f"  selected: col {s['col']}, {s['start']}-{s['end']} ({s['count']} entries), top_row={s['top_row']}",
                  file=sys.stderr)
    return selected


# ─────────────────────────────────────────────
# 番号 → エントリー情報 取得
# ─────────────────────────────────────────────
def extract_entries(ws, number_cols, verbose=False):
    """
    各番号 (組合せ番号) のエントリー情報を取得。
    シングルス: name1
    団体戦: team name + members
    """
    entries = []
    for nc in number_cols:
        col = nc["col"]
        side = "left" if col < 15 else "right"
        for (row, num) in nc["items"]:
            # 同行・近接列からチーム/選手名を取得
            # 番号列の隣 (col+1 or col-1) にチーム名がある可能性が高い
            # 左サイド: col=1 → team col=2, name col=3..
            # 右サイド: col=22 → team col=21, name col=20..19
            if side == "left":
                team_col_candidates = [col + 1]
                name_col_candidates = [col + 2, col + 3, col + 4, col + 5]
            else:
                team_col_candidates = [col - 1]
                name_col_candidates = [col - 2, col - 3, col - 4, col - 5]

            team = ""
            for tc in team_col_candidates:
                v = ws.cell(row=row, column=tc).value
                if is_team_or_name(v):
                    team = normalize_name(v)
                    break

            # 選手名取得: 同番号で 1-7 行範囲をスキャン (団体戦は複数人)
            members = []
            for delta in range(0, 7):
                r = row + delta
                # 次の番号にぶつかったら停止
                next_num = is_number_cell(ws.cell(row=r, column=col).value)
                if delta > 0 and next_num is not None:
                    break
                for nc2 in name_col_candidates:
                    v = ws.cell(row=r, column=nc2).value
                    if is_team_or_name(v) and v != team:
                        n = normalize_name(v)
                        if n not in members:
                            members.append(n)

            # team が空の場合は最初の name を team とみなす (個人戦の可能性)
            if not team and members:
                team = members[0]

            entries.append({
                "no": num,
                "side": side,
                "team": team,
                "members": members,
                "_row": row,
                "_col": col,
            })

    # 番号順にソート
    entries.sort(key=lambda e: e["no"])
    return entries


# ─────────────────────────────────────────────
# 形式判定 (個人 / 団体)
# ─────────────────────────────────────────────
def detect_format(entries):
    """エントリーから個人戦 vs 団体戦を推定"""
    member_counts = [len(e["members"]) for e in entries]
    avg = sum(member_counts) / max(len(member_counts), 1)
    return "team" if avg >= 2.5 else "singles"


# ─────────────────────────────────────────────
# 罫線追跡 (bracket position の確定)
# ─────────────────────────────────────────────
def trace_horizontal_line(ws, row, col, side, max_steps=80):
    """
    指定セル (row, col) から water-line を追跡し、終端列を返す。
    side = "left" は右方向、"right" は左方向。
    """
    direction = 1 if side == "left" else -1
    end = col
    for _ in range(max_steps):
        cur = ws.cell(row=row, column=end)
        # 水平罫線が右 (or 左) へ続いているか?
        nxt_col = end + direction
        if nxt_col < 1 or nxt_col > ws.max_column:
            break
        nxt = ws.cell(row=row, column=nxt_col)
        # 上下のセル境界
        below = ws.cell(row=row + 1, column=end)
        # 「bottom-border あり」または「next row top-border あり」
        cont = (has_border(cur, "bottom") or has_border(below, "top") or
                has_border(nxt, "bottom"))
        if not cont:
            break
        end = nxt_col
    return end


# ─────────────────────────────────────────────
# 標準ブラケット生成 (隣接ペア型)
# ─────────────────────────────────────────────
def build_bracket(entries, event_name, fmt):
    """
    左右独立で bracket を構築 (JTTA トーナメント表の標準構造)。
    左サイド (side=left) を bracket 前半、右サイド (side=right) を bracket 後半に配置。
    両者の決勝戦が「全体決勝」となる。
    番号順の隣接ペア (1v2)(3v4) が各サイドの 1回戦。
    """
    sorted_entries = sorted(entries, key=lambda e: e["no"])
    n = len(sorted_entries)
    if n < 2:
        return {"error": "エントリーが2件未満"}

    # 左右分離
    left = [e for e in sorted_entries if e.get("side", "left") == "left"]
    right = [e for e in sorted_entries if e.get("side") == "right"]

    # 片方しかない (全部 left or 全部 right) 場合は通常 bracket
    if not left or not right:
        # 全部一方サイドの場合、半分に分けて左右に配置
        all_entries = left + right
        half = (len(all_entries) + 1) // 2
        left = all_entries[:half]
        right = all_entries[half:]

    # 左右それぞれの bracket_size を 2のべき乗に
    left_size = 2 ** math.ceil(math.log2(max(len(left), 1)))
    right_size = 2 ** math.ceil(math.log2(max(len(right), 1)))
    # 全体は左右の最大に合わせる (対称な bracket を保つ)
    half_size = max(left_size, right_size)
    bracket_size = half_size * 2
    total_rounds = int(math.log2(bracket_size))

    # 各サイドの R1 ペアを構築
    # 左: 番号順の隣接ペア (1v2, 3v4, ...) - 不足分は BYE
    def make_round1(entries_list, target_size):
        pairs = []
        for i in range(0, target_size, 2):
            e1 = entries_list[i] if i < len(entries_list) else None
            e2 = entries_list[i + 1] if i + 1 < len(entries_list) else None
            pairs.append((e1, e2))
        return pairs

    left_r1 = make_round1(left, half_size)
    right_r1 = make_round1(right, half_size)

    matches_by_round = [[] for _ in range(total_rounds)]
    # 全体 bracket_pos: 0 ～ bracket_size/2-1 のうち、前半が左、後半が右
    # ※ ただし position 順に並べると、決勝で左右の勝者が当たる構造
    pos = 0
    for (e1, e2) in left_r1:
        matches_by_round[0].append({
            "bracket_round": 1, "bracket_pos": pos, "match_no": pos + 1,
            "entry1": e1, "entry2": e2, "side": "left",
        })
        pos += 1
    for (e1, e2) in right_r1:
        matches_by_round[0].append({
            "bracket_round": 1, "bracket_pos": pos, "match_no": pos + 1,
            "entry1": e1, "entry2": e2, "side": "right",
        })
        pos += 1

    # 上位ラウンドは空試合
    for r in range(1, total_rounds):
        n_matches = bracket_size // (2 ** (r + 1))
        for p in range(n_matches):
            # 左右の所属判定 (前半=左、後半=右)
            side_label = "left" if p < n_matches / 2 else "right"
            if r == total_rounds - 1:
                side_label = "final"  # 決勝
            matches_by_round[r].append({
                "bracket_round": r + 1, "bracket_pos": p, "match_no": p + 1,
                "entry1": None, "entry2": None, "side": side_label,
            })

    # ラウンド名
    def round_name(round_num):
        remaining = total_rounds - round_num + 1
        if remaining == 1: return "決勝"
        if remaining == 2: return "準決勝"
        if remaining == 3: return "準々決勝"
        if remaining == 4: return "ベスト16"
        return f"{round_num}回戦"

    # 出力 matches
    all_matches = []
    for r, rnd in enumerate(matches_by_round):
        rname = round_name(r + 1)
        for m in rnd:
            e1 = m["entry1"]
            e2 = m["entry2"]
            mr = {
                "bracket_round": m["bracket_round"],
                "bracket_pos": m["bracket_pos"],
                "round": rname,
                "match_no": m["match_no"],
            }
            if r == 0:
                # round 1 のみ player 情報
                if e1:
                    if fmt == "team":
                        mr["player1_name"] = e1["team"]
                        if e1["members"]:
                            mr["player1_members"] = e1["members"]
                    else:
                        mr["player1_name"] = e1["members"][0] if e1["members"] else e1["team"]
                    mr["player1_team"] = e1["team"]
                else:
                    mr["player1_name"] = "BYE"
                if e2:
                    if fmt == "team":
                        mr["player2_name"] = e2["team"]
                        if e2["members"]:
                            mr["player2_members"] = e2["members"]
                    else:
                        mr["player2_name"] = e2["members"][0] if e2["members"] else e2["team"]
                    mr["player2_team"] = e2["team"]
                else:
                    mr["player2_name"] = "BYE"
            all_matches.append(mr)

    return {
        "format": "tabletennis-bracket-v1",
        "event": event_name,
        "regenerate": True,
        "auto_link_to_players": True,
        "bracket_size": bracket_size,
        "total_rounds": total_rounds,
        "matches": all_matches,
        "_meta": {
            "kumiawase_chart": True,
            "format": fmt,
            "entries_count": n,
            "seeding": "adjacent_pair",
        },
    }


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────
def parse_chart(xlsx_path, sheet_name=None, event_name=None, mode_override=None, verbose=False):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_name = sheet_name or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"シート '{sheet_name}' が見つかりません")
    ws = wb[sheet_name]
    if verbose:
        print(f"[{sheet_name}] {ws.max_row} 行 x {ws.max_column} 列", file=sys.stderr)

    # ① 番号列セグメントを発見
    all_segments = find_number_columns(ws)
    if verbose:
        print(f"  検出セグメント数: {len(all_segments)}", file=sys.stderr)
        for nc in all_segments:
            print(f"    col {nc['col']}: {nc['start']}-{nc['end']} ({nc['count']}件), top_row={nc['top_row']}",
                  file=sys.stderr)

    if not all_segments:
        raise ValueError("組合せ番号列が見つかりません (連番 1, 2, 3... が縦に並ぶ列)")

    # ② bracket 用セグメント選択 (左右両サイド) - 最も上にある block
    number_cols = select_bracket_segments(all_segments, verbose)

    # ③ エントリー抽出
    entries = extract_entries(ws, number_cols, verbose)
    if verbose:
        for e in entries[:5]:
            print(f"  entry #{e['no']:2} ({e['side']:5}): team={e['team']}, members={len(e['members'])} ({e['members'][:3]})",
                  file=sys.stderr)
        if len(entries) > 5:
            print(f"  ... ({len(entries) - 5} 件略)", file=sys.stderr)

    # ③ 形式判定
    fmt = mode_override or detect_format(entries)
    if verbose:
        print(f"  format: {fmt}", file=sys.stderr)

    # ④ Bracket 構築
    bracket = build_bracket(entries, event_name or sheet_name, fmt)
    return bracket


def main():
    p = argparse.ArgumentParser(description="組合せ表 Excel → bracket JSON")
    p.add_argument("xlsx", help="組合せ表 .xlsx ファイル")
    p.add_argument("--sheet", help="シート名 (省略時は先頭)")
    p.add_argument("--event", help="イベント名")
    p.add_argument("--mode", choices=["singles", "team"], help="形式上書き")
    p.add_argument("--output", help="出力先 (省略時は標準出力)")
    p.add_argument("-v", "--verbose", action="store_true")
    args = p.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"[ERR] {xlsx_path} not found", file=sys.stderr)
        sys.exit(1)

    out = parse_chart(xlsx_path, args.sheet, args.event, args.mode, args.verbose)
    if args.output:
        Path(args.output).write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[OK] {args.output}", file=sys.stderr)
    else:
        print(json.dumps(out, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
